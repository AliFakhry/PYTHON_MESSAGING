import openpyxl
from MongoDB import add_database
from MongoDB import find_num
from sendMessage import send_msg
from sendMessage import return_carrier
import os

def iterate_excel(mycol, prev_db, company, curr_names, email, password):

    open_str = os.path.expanduser(f"~/Downloads/GetFeedback/{company}/Invoice.xlsx")
    excel_sheet = openpyxl.load_workbook(open_str)
    sh = excel_sheet.active

    customers_index = False
    start_index = False

    for i in range(1, sh.max_row + 1):
        for j in range(1, sh.max_column + 1):
            if sh.cell(row = i, column= j).value == "Name":
                customers_index = (i+1,j)
                start_index = i+1

    for j in range(start_index, sh.max_row+1):
        cust_name = sh.cell(row = j, column= customers_index[1]).value
        cust_number = find_num(prev_db, cust_name)
        if cust_name == None or cust_number == None:
            continue
        else:
            if cust_name not in curr_names:
                    add_database(cust_name, mycol)
                    print(cust_name, cust_number)
                    try:
                        send_msg(cust_number, return_carrier(cust_number), cust_name, company, email, password)
                    except:
                        continue

    print("\n")
