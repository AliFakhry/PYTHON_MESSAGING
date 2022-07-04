from pymongo import MongoClient
import os
import openpyxl
from MongoDB import add_database

def main():
    print("PARSE INITIAL. ")
    password = input("ENTER MongoDB Password: ")
    cluster = f"mongodb+srv://Phone_Data:{password}@cluster0.xzech.mongodb.net/?retryWrites=true&w=majority"
    client = MongoClient(cluster)
    mydb = client["Names"]

    company = input("")
    open_str = os.path.expanduser(f"~/Downloads/GetFeedback/{company}_Initial/Customers.xlsx")
    excel_sheet = openpyxl.load_workbook(open_str)
    sh = excel_sheet.active
    mycol = mydb[f"{company}_Initial"]

    customers_index = False
    numebrs_index = False
    start_index = False

    for i in range(1, sh.max_row + 1):
        for j in range(1, sh.max_column + 1):
            if sh.cell(row = i, column= j).value == "Customer":
                customers_index = (i+1,j)
                start_index = i+1
            if sh.cell(row = i, column= j).value == "Phone Numbers":
                numebrs_index = (i+1, j)
                start_index = i+1

    for j in range(start_index, sh.max_row+1):
        cust_name = sh.cell(row = j, column= customers_index[1]).value
        cust_number = (sh.cell(row = j, column=numebrs_index[1]).value)
        if cust_name == None or cust_number == None:
            continue
        else:
            new_str = ""
            for i in cust_number:
                if i.isdigit():
                    new_str = new_str + i
            cust_total = (cust_name, new_str)
            try:
                add_database(cust_total, mycol)
            except:
                continue

if __name__ == "__main__":
    main()